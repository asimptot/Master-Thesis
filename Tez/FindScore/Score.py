from  openpyxl import *
import xlrd

def calculate (id, isci_sayisi):
    worker = 1
    j = 2
    loc = (r'D:\Projects\Tez\FindScore\Scores.xlsx')
    wb = load_workbook(r'D:\Projects\Tez\FindScore\Scores.xlsx')
    wb2 = xlrd.open_workbook(loc)

    sheets = wb.sheetnames
    sheet1 = wb[sheets[id]]
    sheet2 = wb2.sheet_by_index(0)

    for i in range(1, worker + 1):
        temp = sheet2.cell_value(i, 0)
        print("\n", "for ", temp)
        fixed_showstopper = input("Enter fixed showstopper bug numbers: ")
        puan = int(fixed_showstopper) * 20
        sheet1.cell(row=i + 1, column=j).value = puan

        fixed_high = input("Enter fixed high bug numbers: ")
        puan2 = int(fixed_high) * 10
        sheet1.cell(row=i + 1, column=j + 1).value = puan2

        fixed_medium = input("Enter fixed medium bug numbers: ")
        puan3 = int(fixed_medium) * 4
        sheet1.cell(row=i + 1, column=j + 2).value = puan3

        fixed_low = input("Enter fixed low bug numbers: ")
        puan4 = int(fixed_low) * 2
        sheet1.cell(row=i + 1, column=j + 3).value = puan4

        retest_showstopper = input("Enter retest showstopper bug numbers: ")
        puan5 = int(retest_showstopper) * -20
        sheet1.cell(row=i + 1, column=j + 4).value = puan5

        retest_high = input("Enter retest high bug numbers: ")
        puan6 = int(retest_high) * -10
        sheet1.cell(row=i + 1, column=j + 5).value = puan6

        retest_medium = input("Enter retest medium bug numbers: ")
        puan7 = int(retest_medium) * -4
        sheet1.cell(row=i + 1, column=j + 6).value = puan7

        retest_low = input("Enter retest low bug numbers: ")
        puan8 = int(retest_low) * -2
        sheet1.cell(row=i + 1, column=j + 7).value = puan8

        qa_showstopper = input("Enter fixed showstopper bug numbers by Quality Assurance: ")
        puan9 = int(qa_showstopper) * -20
        sheet1.cell(row=i + 1, column=j + 8).value = puan9

        qa_high = input("Enter fixed high bug numbers by Quality Assurance: ")
        puan10 = int(qa_high) * -10
        sheet1.cell(row=i + 1, column=j + 9).value = puan10

        qa_medium = input("Enter fixed medium bug numbers by Quality Assurance: ")
        puan11 = int(qa_medium) * -4
        sheet1.cell(row=i + 1, column=j + 10).value = puan11

        qa_low = input("Enter fixed low bug numbers by Quality Assurance: ")
        puan12 = int(qa_low) * -2
        sheet1.cell(row=i + 1, column=j + 11).value = puan12

        sure = input("How long did the test take? Enter data in hours: ")
        puan13 = 1 / int(sure) * 20
        sheet1.cell(row=i + 1, column=j + 12).value = int(puan13)

        toplam = puan + puan2 + puan3 + puan4 + puan5 + puan6 + puan7 + puan8 + puan9 + puan10 + puan11 + puan12 + puan13
        sheet1.cell(row=i + 1, column=j + 13).value = int(toplam)
        wb.save('Scores.xlsx')

        j = 2

print("\nEnter datas for Release Approval Test:")
calculate(0, 1)

print("\nEnter datas for Portal Test:")
calculate(1, 1)

print("\nEnter datas for Logging Test:")
calculate(2, 1)
